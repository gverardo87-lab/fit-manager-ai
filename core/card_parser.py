"""
CardParser - Import and parse workout cards from Excel/Word files

Parses Chiara's existing workout cards into structured data.
Handles hybrid documents: mixed tables + free text, circuit formats,
Italian notation, non-standard set/rep schemes.

Designed around real trainer documents, not clean spreadsheets.
"""

import io
import re
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import List, Optional, Tuple

from core.exercise_archive import ExerciseArchive


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class ParsedExercise:
    """Single exercise extracted from a workout card."""
    name: str
    canonical_id: Optional[str] = None  # matched ID from exercise_db
    match_score: float = 0.0            # fuzzy match confidence
    sets: Optional[int] = None
    reps: Optional[str] = None          # "8-12" or "3x10" or "12"
    rest_seconds: Optional[int] = None
    load_note: Optional[str] = None     # "70%", "60kg", "RM"
    notes: Optional[str] = None
    day_section: Optional[str] = None   # "GIORNO 1", "GIORNO 2" etc.
    row_index: int = 0


@dataclass
class ParsedCardMetadata:
    """Metadata detected from a workout card."""
    detected_goal: Optional[str] = None
    detected_split: Optional[str] = None
    detected_weeks: Optional[int] = None
    detected_sessions_per_week: Optional[int] = None
    trainer_notes: List[str] = field(default_factory=list)
    sheet_names: List[str] = field(default_factory=list)
    client_name: Optional[str] = None
    days_found: List[str] = field(default_factory=list)


@dataclass
class ParsedCard:
    """Complete result of parsing a workout card file."""
    raw_text: str
    exercises: List[ParsedExercise]
    metadata: ParsedCardMetadata
    parse_confidence: float = 0.0       # 0.0-1.0 overall confidence


# ============================================================================
# COLUMN DETECTION PATTERNS (expanded for Chiara's format)
# ============================================================================

EXERCISE_HEADERS = [
    "esercizio", "exercise", "nome", "name", "movimento", "movement",
    "es.", "ex.", "esercizi",
]

# Combined sets+reps headers (Chiara uses these)
SETS_REPS_HEADERS = [
    "serie-ripetizioni", "serie/ripetizioni", "serie - ripetizioni",
    "serie/reps", "serie e ripetizioni", "sets/reps", "sets-reps",
]

SETS_HEADERS = [
    "serie", "sets", "set", "s", "ser",
]
REPS_HEADERS = [
    "ripetizioni", "reps", "rep", "r", "rip", "ripet",
]
REST_HEADERS = [
    "recupero", "rest", "pausa", "rec", "rec.",
]
LOAD_HEADERS = [
    "carico", "load", "peso", "weight", "kg", "%", "intensita", "intensity",
    "sovraccarichi", "sovraccarico",
]
NOTE_HEADERS = [
    "note", "notes", "nota", "annotazioni", "commenti", "tempo",
]

# Goal detection keywords
GOAL_KEYWORDS = {
    "strength": ["forza", "strength", "1rm", "massimale", "max", "powerlifting"],
    "hypertrophy": ["ipertrofia", "hypertrophy", "massa", "muscle", "volume", "bodybuilding"],
    "fat_loss": ["dimagrimento", "fat loss", "definizione", "cut", "lean"],
    "endurance": ["resistenza", "endurance", "stamina", "conditioning"],
    "functional": ["funzionale", "functional", "mobilita", "mobility", "calisthenics",
                    "circuito", "circuit", "pha", "total body"],
}

# Split detection keywords
SPLIT_KEYWORDS = {
    "full_body": ["full body", "total body", "tutto corpo", "completo"],
    "upper_lower": ["upper", "lower", "alto", "basso", "superiore", "inferiore"],
    "push_pull_legs": ["push", "pull", "legs", "ppl", "spinta", "tirata", "gambe"],
    "bro_split": ["petto", "schiena", "spalle", "braccia", "chest", "back", "arms"],
    "circuit": ["circuito", "circuit", "pha", "tabata", "hiit", "emom"],
}

# Section/title words to skip as exercise names
SKIP_WORDS = {
    "esercizio", "exercise", "serie", "reps", "ripetizioni", "recupero", "carico",
    "note", "circuito", "riscaldamento", "riscaldamento generale",
    "riscaldamento specifico", "circuito riscaldamento", "defaticamento",
    "allungamento", "note allenamento", "addome", "addome isometrico",
    "richiamo", "circuito ritorno venoso", "total body", "sovraccarichi",
    "serie-ripetizioni", "serie/ripetizioni", "sovraccarico",
    "lunedi", "martedi", "mercoledi", "giovedi", "venerdi", "sabato", "domenica",
    "rest", "day", "esempio", "esempio svolgimento", "buon lavoro",
    "misure", "peso", "vita", "fianchi", "coscia", "braccio", "seno",
    "start", "inizio", "fine", "nome",
    "allenamento", "allenamento 1", "allenamento 2", "allenamento 3",
}

# Patterns that indicate a row is an exercise (even without DB match)
EXERCISE_INDICATORS = re.compile(
    r'\b(squat|press|curl|row|push|pull|plank|crunch|lunge|deadlift|'
    r'stacco|panca|affondi|alzate|rematore|tirate|trazioni|dip|'
    r'frog|swing|thruster|clean|jerk|snatch|hip.?thrust|'
    r'lat.?machine|vertical|leg.?press|leg.?curl|leg.?extension|'
    r'calf|shoulder|chest|fly|croci|military|'
    r'v.?up|mountain|climber|jumping|jack|burpee|'
    r'sumo|goblet|romanian|bulgarian|glute|bridge|'
    r'push.?up|push.?down|pull.?up|chin.?up|'
    r'traction|femorale|arco|iperestensioni|'
    r'tricipiti|bicipiti|addominale)\b',
    re.IGNORECASE
)

# Regex for sets x reps patterns in Italian trainer notation
SETS_REPS_PATTERN = re.compile(
    r'(\d+)\s*[xX×]\s*(\d+(?:[/-]\d+)?(?:\s*x\s*gamba)?)',
    re.IGNORECASE
)

# Time-based work pattern (circuits)
TIME_WORK_PATTERN = re.compile(
    r"(\d+)['\"]?\s*(?:on|work|lavoro)\s*(\d+)['\"]?\s*(?:off|rest|pausa)",
    re.IGNORECASE
)

# Reps-only pattern
REPS_ONLY_PATTERN = re.compile(r'^(\d+)$')

# "Giri" (rounds) pattern
GIRI_PATTERN = re.compile(r'(\d+)\s*(?:giri|rounds?|giro)', re.IGNORECASE)

# Day section pattern
DAY_PATTERN = re.compile(
    r'(?:GIORNO|DAY|JOUR)\s*(\d+)',
    re.IGNORECASE
)

# Rest time pattern from text
REST_PATTERN = re.compile(
    r"(\d+)['\"]?\s*(?:sec(?:ondi)?|s\b|\"|'')|"
    r"(\d+)[':]\s*(\d+)['\"]?\s*(?:min)?|"
    r"(\d+)\s*(?:minut[io]|min|')",
    re.IGNORECASE
)


class CardParser:
    """Parses workout cards from Excel and Word files."""

    def __init__(self):
        self._exercise_names = {}
        archive = ExerciseArchive()
        for ex in archive.get_all():
            name = ex.get('name', '').lower()
            italian = (ex.get('italian_name') or '').lower()
            ex_id = str(ex.get('id', ''))
            if name:
                self._exercise_names[name] = ex_id
            if italian:
                self._exercise_names[italian] = ex_id

    def parse_file(self, file_bytes: bytes, filename: str) -> ParsedCard:
        """Auto-detect file format and parse."""
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

        if ext in ("xlsx", "xls"):
            return self.parse_excel(file_bytes, filename)
        elif ext in ("docx", "doc"):
            return self.parse_word(file_bytes, filename)
        else:
            return ParsedCard(
                raw_text="",
                exercises=[],
                metadata=ParsedCardMetadata(),
                parse_confidence=0.0,
            )

    # ================================================================
    # EXCEL PARSING
    # ================================================================

    def parse_excel(self, file_bytes: bytes, filename: str) -> ParsedCard:
        """Parse an Excel workout card."""
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        all_exercises: List[ParsedExercise] = []
        raw_lines: List[str] = []
        metadata = ParsedCardMetadata(sheet_names=wb.sheetnames)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            col_map = self._detect_columns(rows)
            start_row = 1 if col_map["header_row"] == 0 else col_map["header_row"] + 1
            for i, row in enumerate(rows[start_row:], start=start_row):
                line = " | ".join(str(c) for c in row if c is not None)
                raw_lines.append(line)

                ex = self._parse_exercise_row(row, col_map, i)
                if ex:
                    all_exercises.append(ex)

            all_text = " ".join(
                str(cell) for row in rows for cell in row if cell is not None
            ).lower()
            self._detect_metadata(all_text, metadata)

        for ex in all_exercises:
            canonical_id, score = self._normalize_exercise_name(ex.name)
            ex.canonical_id = canonical_id
            ex.match_score = score

        confidence = self._calculate_confidence(all_exercises)

        return ParsedCard(
            raw_text="\n".join(raw_lines),
            exercises=all_exercises,
            metadata=metadata,
            parse_confidence=confidence,
        )

    # ================================================================
    # WORD PARSING (redesigned for hybrid documents)
    # ================================================================

    def parse_word(self, file_bytes: bytes, filename: str) -> ParsedCard:
        """
        Parse a Word workout card with hybrid format.

        Strategy: process ALL content in document order (paragraphs + tables),
        detect day sections, extract exercises from both tables and free text.
        """
        try:
            import docx
        except ImportError:
            return ParsedCard(
                raw_text="python-docx non installato",
                exercises=[],
                metadata=ParsedCardMetadata(),
                parse_confidence=0.0,
            )

        doc = docx.Document(io.BytesIO(file_bytes))
        all_exercises: List[ParsedExercise] = []
        raw_lines: List[str] = []
        metadata = ParsedCardMetadata()
        current_day = None
        all_text_parts = []
        row_counter = 0

        # Process document in order: iterate body elements
        for element in doc.element.body:
            tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag

            if tag == 'p':
                # Paragraph
                text = element.text or ''
                # Also get text from runs (handles formatted text)
                if not text:
                    text = ''.join(
                        node.text or '' for node in element.iter()
                        if node.text
                    )
                text = text.strip()

                if not text:
                    continue

                raw_lines.append(text)
                all_text_parts.append(text.lower())

                # Check for day section marker
                day_match = DAY_PATTERN.search(text)
                if day_match:
                    current_day = f"GIORNO {day_match.group(1)}"
                    if current_day not in metadata.days_found:
                        metadata.days_found.append(current_day)
                    continue

                # Check for client name (first line pattern "NOME: ...")
                if text.upper().startswith("NOME:"):
                    metadata.client_name = text.split(":", 1)[1].strip()
                    continue

                # Try to extract exercises from free text (bullet lists, inline)
                free_exercises = self._parse_free_text_exercises(text, row_counter)
                for ex in free_exercises:
                    ex.day_section = current_day
                    all_exercises.append(ex)
                    row_counter += 1

                # Collect trainer notes (longer descriptive text)
                if len(text) > 30 and not self._is_exercise_text(text):
                    if not any(skip in text.lower() for skip in [
                        "buon lavoro", "divertiti", "non mollare",
                        "capolavoro", "nave da crociera", "buon vento",
                    ]):
                        metadata.trainer_notes.append(text)

            elif tag == 'tbl':
                # Table - find the corresponding docx Table object
                table = self._find_table_for_element(doc, element)
                if table is None:
                    continue

                table_exercises = self._parse_word_table(
                    table, row_counter, current_day
                )
                all_exercises.extend(table_exercises)
                row_counter += len(table_exercises)

                # Also add raw text from table
                for tr in table.rows:
                    row_text = " | ".join(cell.text.strip() for cell in tr.cells)
                    raw_lines.append(row_text)
                    all_text_parts.append(row_text.lower())

        # Detect metadata from all text
        all_text = " ".join(all_text_parts)
        self._detect_metadata(all_text, metadata)

        # Detect sessions from days found
        if metadata.days_found and not metadata.detected_sessions_per_week:
            metadata.detected_sessions_per_week = len(metadata.days_found)

        # Normalize exercise names and filter duplicates
        seen_names = set()
        unique_exercises = []
        for ex in all_exercises:
            # Clean exercise name
            ex.name = self._clean_exercise_name(ex.name)
            if not ex.name or len(ex.name) < 2:
                continue

            # Skip duplicates (same name in same day)
            key = (ex.name.lower(), ex.day_section)
            if key in seen_names:
                continue
            seen_names.add(key)

            canonical_id, score = self._normalize_exercise_name(ex.name)
            ex.canonical_id = canonical_id
            ex.match_score = score
            unique_exercises.append(ex)

        confidence = self._calculate_confidence(unique_exercises)

        return ParsedCard(
            raw_text="\n".join(raw_lines),
            exercises=unique_exercises,
            metadata=metadata,
            parse_confidence=confidence,
        )

    def _find_table_for_element(self, doc, tbl_element):
        """Find the docx.table.Table object matching an XML element."""
        for table in doc.tables:
            if table._tbl is tbl_element:
                return table
        return None

    def _parse_word_table(
        self, table, start_row: int, current_day: Optional[str]
    ) -> List[ParsedExercise]:
        """
        Parse a Word table into exercises.

        Handles Chiara's format: exercise tables have 3-4 columns,
        while schedule/measurement tables are different.
        """
        rows = []
        for tr in table.rows:
            row_data = tuple(cell.text.strip() for cell in tr.cells)
            rows.append(row_data)

        if not rows or len(rows) < 2:
            return []

        # Classify table: is it an exercise table?
        if not self._is_exercise_table(rows):
            return []

        # Detect columns with expanded header matching
        col_map = self._detect_columns_v2(rows)

        exercises = []
        start = col_map["header_row"] + 1 if col_map["header_row"] >= 0 else 0

        for i, row in enumerate(rows[start:], start=start):
            ex = self._parse_exercise_row_v2(row, col_map, start_row + i)
            if ex:
                # Check if this row is a day marker within the table
                day_match = DAY_PATTERN.search(ex.name)
                if day_match:
                    current_day = f"GIORNO {day_match.group(1)}"
                    continue

                ex.day_section = current_day
                exercises.append(ex)

        return exercises

    def _is_exercise_table(self, rows: list) -> bool:
        """
        Determine if a table contains exercises vs schedule/measurements.

        Exercise tables typically have:
        - A header with "esercizio" or "serie" or "recupero"
        - Or rows that contain exercise-like text
        """
        all_text = " ".join(
            " ".join(str(c) for c in row) for row in rows
        ).lower()

        # Check for exercise table headers
        exercise_headers = (
            EXERCISE_HEADERS + SETS_REPS_HEADERS + SETS_HEADERS +
            REST_HEADERS + LOAD_HEADERS
        )
        for header in exercise_headers:
            if header in all_text:
                return True

        # Check for exercise indicator words
        if EXERCISE_INDICATORS.search(all_text):
            return True

        # Check for sets/reps patterns
        if SETS_REPS_PATTERN.search(all_text):
            return True

        # Skip schedule tables (days of the week in header)
        first_row = " ".join(str(c) for c in rows[0]).lower()
        weekdays = ["lunedi", "martedi", "mercoledi", "giovedi", "venerdi", "sabato"]
        if sum(1 for wd in weekdays if wd in first_row) >= 3:
            return False

        # Skip measurement tables
        measurements = ["vita", "fianchi", "coscia", "braccio", "seno"]
        if sum(1 for m in measurements if m in first_row) >= 2:
            return False

        return False

    def _detect_columns_v2(self, rows: list) -> dict:
        """
        Enhanced column detection for Chiara's varied formats.

        Handles combined "SERIE-RIPETIZIONI" columns and
        non-standard header names like "SOVRACCARICHI".
        """
        col_map = {
            "exercise": None,
            "sets_reps": None,  # combined column
            "sets": None,
            "reps": None,
            "rest": None,
            "load": None,
            "notes": None,
            "header_row": -1,
        }

        def _header_matches(cell_text: str, header_list: list) -> bool:
            """Match cell text against header list.
            Short headers (1-2 chars) require exact match.
            Longer headers allow startswith match.
            """
            for h in header_list:
                if h == cell_text:
                    return True
                if len(h) >= 3 and cell_text.startswith(h):
                    return True
            return False

        # Scan first rows for headers
        for row_idx, row in enumerate(rows[:5]):
            if row is None:
                continue
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip().lower()
                cell_clean = re.sub(r'[^a-z/\-\s]', '', cell_str)

                # Combined sets+reps (Chiara's "SERIE-RIPETIZIONI")
                if any(h in cell_clean for h in SETS_REPS_HEADERS):
                    col_map["sets_reps"] = col_idx
                    col_map["header_row"] = row_idx
                elif _header_matches(cell_clean, EXERCISE_HEADERS):
                    col_map["exercise"] = col_idx
                    col_map["header_row"] = row_idx
                elif _header_matches(cell_clean, REST_HEADERS):
                    col_map["rest"] = col_idx
                elif _header_matches(cell_clean, LOAD_HEADERS):
                    col_map["load"] = col_idx
                elif _header_matches(cell_clean, SETS_HEADERS):
                    col_map["sets"] = col_idx
                elif _header_matches(cell_clean, REPS_HEADERS):
                    col_map["reps"] = col_idx
                elif _header_matches(cell_clean, NOTE_HEADERS):
                    col_map["notes"] = col_idx

        # Fallback: if we found a header row but no exercise column,
        # assume first column is exercise
        if col_map["header_row"] >= 0 and col_map["exercise"] is None:
            col_map["exercise"] = 0

        # Fallback: if no headers found at all, try to detect from content
        if col_map["header_row"] < 0:
            col_map = self._detect_columns_from_content(rows, col_map)

        return col_map

    def _detect_columns_from_content(self, rows: list, col_map: dict) -> dict:
        """Fallback: detect columns by analyzing cell content patterns."""
        if not rows or len(rows) < 2:
            return col_map

        # Find the first column with text that looks like exercise names
        for row in rows[:6]:
            if row is None:
                continue
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                cell_lower = cell_str.lower()

                # Skip if it's a skip word
                if cell_lower in SKIP_WORDS:
                    continue

                # Is this cell an exercise name?
                if (len(cell_str) > 3 and
                    EXERCISE_INDICATORS.search(cell_str) and
                    col_map["exercise"] is None):
                    col_map["exercise"] = col_idx
                    col_map["header_row"] = -1  # no header row
                    break

            if col_map["exercise"] is not None:
                break

        # Try to identify other columns by content
        if col_map["exercise"] is not None:
            ex_col = col_map["exercise"]
            for row in rows[1:5]:
                if row is None:
                    continue
                for offset in range(1, min(5, len(row))):
                    col_idx = (ex_col + offset) % len(row)
                    if col_idx == ex_col:
                        continue
                    val = row[col_idx]
                    if val is None:
                        continue
                    val_str = str(val).strip()

                    # Sets/reps pattern
                    if SETS_REPS_PATTERN.search(val_str) and col_map["sets_reps"] is None:
                        col_map["sets_reps"] = col_idx
                    # Rest pattern
                    elif REST_PATTERN.search(val_str) and col_map["rest"] is None:
                        col_map["rest"] = col_idx

        return col_map

    def _parse_exercise_row_v2(
        self, row: tuple, col_map: dict, row_idx: int
    ) -> Optional[ParsedExercise]:
        """
        Enhanced row parsing for Chiara's formats.

        Handles combined sets/reps columns, inline sets notation,
        and various Italian workout notation styles.
        """
        ex_col = col_map.get("exercise", 0) or 0
        if ex_col >= len(row):
            return None

        cell_val = row[ex_col]
        if cell_val is None:
            return None

        name = str(cell_val).strip()

        # Skip empty, short, or header/section names
        if len(name) < 2:
            return None
        name_lower = name.lower().strip()
        if name_lower in SKIP_WORDS or name_lower in ("none", "-", "", "//"):
            return None

        # Skip if it looks like a pure number or date
        if re.match(r'^\d+[°^]?\s*(sett|settimana)?$', name_lower):
            return None

        # Skip weekday names
        if name_lower.rstrip("'") in (
            "lunedi", "martedi", "mercoledi", "giovedi", "venerdi", "sabato", "domenica"
        ):
            return None

        # Skip note-like long sentences (> 80 chars usually aren't exercise names)
        if len(name) > 80:
            return None

        ex = ParsedExercise(name=name, row_index=row_idx)

        # Parse combined sets/reps column
        if col_map.get("sets_reps") is not None:
            sr_col = col_map["sets_reps"]
            if sr_col < len(row) and row[sr_col]:
                self._parse_sets_reps_combined(str(row[sr_col]), ex)

        # Parse separate sets column
        if col_map.get("sets") is not None and col_map["sets"] < len(row):
            val = row[col_map["sets"]]
            if val is not None:
                try:
                    ex.sets = int(float(str(val)))
                except (ValueError, TypeError):
                    # Try to extract from text like "4x12"
                    self._parse_sets_reps_combined(str(val), ex)

        # Parse separate reps column
        if col_map.get("reps") is not None and col_map["reps"] < len(row):
            val = row[col_map["reps"]]
            if val is not None and ex.reps is None:
                ex.reps = str(val).strip()

        # Parse rest column
        if col_map.get("rest") is not None and col_map["rest"] < len(row):
            val = row[col_map["rest"]]
            if val is not None:
                ex.rest_seconds = self._parse_rest_time(str(val))

        # Parse load column
        if col_map.get("load") is not None and col_map["load"] < len(row):
            val = row[col_map["load"]]
            if val is not None:
                load_str = str(val).strip()
                if load_str and load_str not in ("-", "//", "none"):
                    ex.load_note = load_str

        # If no sets/reps found from columns, try to extract from name itself
        # e.g. "V-UP  20" or "PLANK 20""
        if ex.sets is None and ex.reps is None:
            self._extract_inline_reps(ex)

        return ex

    def _parse_sets_reps_combined(self, text: str, ex: ParsedExercise):
        """
        Parse combined sets/reps text in various Italian formats.

        Handles: "4x12", "4x12 x gamba", "40"ON 20"OFF 3 GIRI",
        "20", "SUPERSET PIRAMIDALE 18-18/2-2", "20+20+20 4 GIRI"
        """
        text = text.strip()
        if not text or text in ("-", "//"):
            return

        # Standard sets x reps: "4x12", "5x8/10", "4x12 x gamba"
        m = SETS_REPS_PATTERN.search(text)
        if m:
            ex.sets = int(m.group(1))
            ex.reps = m.group(2)
            return

        # Time-based circuit: "40"ON 20"OFF"
        m = TIME_WORK_PATTERN.search(text)
        if m:
            ex.reps = f"{m.group(1)}\"ON {m.group(2)}\"OFF"
            # Check for rounds
            giri = GIRI_PATTERN.search(text)
            if giri:
                ex.sets = int(giri.group(1))
            return

        # Rounds/giri with reps: "20+20+20 4 GIRI" or "3 GIRI"
        giri = GIRI_PATTERN.search(text)
        if giri:
            ex.sets = int(giri.group(1))
            # Extract reps part (everything before "giri")
            before_giri = text[:giri.start()].strip()
            if before_giri and before_giri not in ("-", "//"):
                ex.reps = before_giri
            return

        # Plain number = reps
        m = REPS_ONLY_PATTERN.match(text)
        if m:
            ex.reps = text
            return

        # Seconds-based (for holds): "20"", "35""
        if re.match(r'^\d+["\']', text):
            ex.reps = text
            return

        # Anything else: store as reps text
        if len(text) < 30:
            ex.reps = text

    def _extract_inline_reps(self, ex: ParsedExercise):
        """Extract reps from exercise name when they're inline, e.g. 'V-UP  20'."""
        # Pattern: "EXERCISE_NAME  NUMBER" at end
        m = re.search(r'\s+(\d+)\s*$', ex.name)
        if m:
            reps_val = int(m.group(1))
            if 1 <= reps_val <= 100:
                ex.reps = str(reps_val)
                ex.name = ex.name[:m.start()].strip()

        # Pattern: "EXERCISE_NAME  NUMBERxNUMBER"
        m = re.search(r'\s+(\d+)\s*[xX×]\s*(\d+(?:[/-]\d+)?)\s*$', ex.name)
        if m:
            ex.sets = int(m.group(1))
            ex.reps = m.group(2)
            ex.name = ex.name[:m.start()].strip()

    def _parse_rest_time(self, text: str) -> Optional[int]:
        """Parse rest time from various formats."""
        text = text.strip().lower()
        if not text or text in ("-", "//", "none"):
            return None

        # "1'30"" or "1:30"
        m = re.search(r'(\d+)[\':](\d+)', text)
        if m:
            return int(m.group(1)) * 60 + int(m.group(2))

        # "90s", "90 sec", "90""
        m = re.search(r'(\d+)\s*(?:s(?:ec(?:ondi)?)?|"|\'\')', text)
        if m:
            return int(m.group(1))

        # "2'" or "2 min"
        m = re.search(r'(\d+)\s*(?:\'|min)', text)
        if m:
            return int(m.group(1)) * 60

        # Plain number (assume seconds if < 300, minutes if >= 300)
        m = re.match(r'^(\d+)$', text)
        if m:
            val = int(m.group(1))
            return val if val < 300 else val

        return None

    def _parse_free_text_exercises(
        self, text: str, start_row: int
    ) -> List[ParsedExercise]:
        """
        Extract exercises from free text paragraphs.

        Handles bullet lists like "• 10 squat" and inline formats.
        """
        exercises = []
        text = text.strip()

        # Skip very long paragraphs (likely descriptive text, not exercises)
        if len(text) > 150:
            return []

        # Skip known non-exercise patterns
        text_lower = text.lower()
        if any(skip in text_lower for skip in [
            "ricordati", "ricorda", "esempio", "n.b.", "nb:",
            "importante", "fondamentale seguire", "alimentazione",
            "buon lavoro", "divertiti", "obbiettivo", "obiettivo",
            "dovrai", "ogni sera", "la mattina", "nei giorni",
            "oltre all'allenamento", "tapis roulant",
            "allungamento finali", "stretching",
            "al mattino", "volte a settimana",
        ]):
            return []

        # Bullet list items: "• 10 squat" or "• Mobilità anche"
        bullet_pattern = re.compile(
            r'[•\-\*]\s*(\d+)?\s*(.+?)(?:\s*$)',
        )

        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                continue

            m = bullet_pattern.match(line)
            if m:
                reps_str = m.group(1)
                name = m.group(2).strip()

                if len(name) < 3:
                    continue
                if name.lower() in SKIP_WORDS:
                    continue

                # Only include if it looks like an exercise
                if EXERCISE_INDICATORS.search(name) or reps_str:
                    ex = ParsedExercise(
                        name=name,
                        reps=reps_str,
                        row_index=start_row,
                    )
                    exercises.append(ex)

        return exercises

    def _is_exercise_text(self, text: str) -> bool:
        """Check if text contains exercise-related content."""
        return bool(EXERCISE_INDICATORS.search(text))

    def _clean_exercise_name(self, name: str) -> str:
        """Clean up exercise name: remove extra whitespace, trailing numbers etc."""
        # Remove multiple spaces
        name = re.sub(r'\s+', ' ', name).strip()
        # Remove leading/trailing special chars
        name = name.strip('•-*·')
        return name.strip()

    # ================================================================
    # SHARED HELPERS (column detection v1 for Excel)
    # ================================================================

    def _detect_columns(self, rows: list) -> dict:
        """Detect columns - used by Excel parser."""
        col_map = {
            "exercise": None,
            "sets_reps": None,
            "sets": None,
            "reps": None,
            "rest": None,
            "load": None,
            "notes": None,
            "header_row": 0,
        }

        def _header_match(cell_text: str, header_list: list) -> bool:
            for h in header_list:
                if h == cell_text:
                    return True
                if len(h) >= 3 and cell_text.startswith(h):
                    return True
            return False

        for row_idx, row in enumerate(rows[:5]):
            if row is None:
                continue
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip().lower()

                # Check combined sets/reps first
                if any(h in cell_str for h in SETS_REPS_HEADERS):
                    col_map["sets_reps"] = col_idx
                    col_map["header_row"] = row_idx
                elif _header_match(cell_str, EXERCISE_HEADERS):
                    col_map["exercise"] = col_idx
                    col_map["header_row"] = row_idx
                elif _header_match(cell_str, REST_HEADERS):
                    col_map["rest"] = col_idx
                elif _header_match(cell_str, LOAD_HEADERS):
                    col_map["load"] = col_idx
                elif _header_match(cell_str, SETS_HEADERS):
                    col_map["sets"] = col_idx
                elif _header_match(cell_str, REPS_HEADERS):
                    col_map["reps"] = col_idx
                elif _header_match(cell_str, NOTE_HEADERS):
                    col_map["notes"] = col_idx

        # Fallback
        if col_map["exercise"] is None and rows:
            for row in rows[1:4]:
                if row is None:
                    continue
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str) and len(cell) > 3:
                        col_map["exercise"] = col_idx
                        break
                if col_map["exercise"] is not None:
                    break

        return col_map

    def _parse_exercise_row(
        self, row: tuple, col_map: dict, row_idx: int
    ) -> Optional[ParsedExercise]:
        """Parse a single row - used by Excel parser."""
        ex_col = col_map.get("exercise")
        if ex_col is None:
            return None
        if ex_col >= len(row) or row[ex_col] is None:
            return None

        name = str(row[ex_col]).strip()
        if len(name) < 2 or name.lower() in ("none", "-", ""):
            return None
        if name.lower() in EXERCISE_HEADERS or name.lower() in SKIP_WORDS:
            return None

        ex = ParsedExercise(name=name, row_index=row_idx)

        # Combined sets/reps
        if col_map.get("sets_reps") is not None and col_map["sets_reps"] < len(row):
            val = row[col_map["sets_reps"]]
            if val is not None:
                self._parse_sets_reps_combined(str(val), ex)

        # Separate sets
        if col_map.get("sets") is not None and col_map["sets"] < len(row):
            val = row[col_map["sets"]]
            if val is not None and ex.sets is None:
                try:
                    ex.sets = int(float(str(val)))
                except (ValueError, TypeError):
                    pass

        # Separate reps
        if col_map.get("reps") is not None and col_map["reps"] < len(row):
            val = row[col_map["reps"]]
            if val is not None and ex.reps is None:
                ex.reps = str(val).strip()

        # Rest
        if col_map.get("rest") is not None and col_map["rest"] < len(row):
            val = row[col_map["rest"]]
            if val is not None:
                ex.rest_seconds = self._parse_rest_time(str(val))

        # Load
        if col_map.get("load") is not None and col_map["load"] < len(row):
            val = row[col_map["load"]]
            if val is not None:
                load_str = str(val).strip()
                if load_str and load_str not in ("-", "//"):
                    ex.load_note = load_str

        # Notes
        if col_map.get("notes") is not None and col_map["notes"] < len(row):
            val = row[col_map["notes"]]
            if val is not None:
                ex.notes = str(val).strip()

        return ex

    # ================================================================
    # EXERCISE NAME MATCHING
    # ================================================================

    def _normalize_exercise_name(self, name: str) -> Tuple[str, float]:
        """
        Match exercise name against exercise database using fuzzy matching.

        Lower threshold (0.45) for Chiara's informal Italian names.
        Also tries partial matching for compound names like
        "AFFONDI POSTERIORI + ALZATE FRONTALI".
        """
        name_lower = name.lower().strip()

        # Exact match
        if name_lower in self._exercise_names:
            return self._exercise_names[name_lower], 1.0

        # Try each part if it's a compound exercise (A + B)
        best_id = None
        best_score = 0.0

        parts = re.split(r'\s*[+&]\s*', name_lower)
        for part in parts:
            part = part.strip()
            if not part:
                continue

            # Exact match on part
            if part in self._exercise_names:
                return self._exercise_names[part], 0.95

            # Fuzzy match on part
            for db_name, ex_id in self._exercise_names.items():
                score = SequenceMatcher(None, part, db_name).ratio()
                if score > best_score:
                    best_score = score
                    best_id = ex_id

        # Also fuzzy match the full name
        for db_name, ex_id in self._exercise_names.items():
            score = SequenceMatcher(None, name_lower, db_name).ratio()
            if score > best_score:
                best_score = score
                best_id = ex_id

        # Lower threshold for real trainer documents
        if best_score >= 0.45:
            return best_id, round(best_score, 2)

        # Even with no DB match, if it looks like an exercise, give partial score
        if EXERCISE_INDICATORS.search(name):
            return None, 0.3

        return None, 0.0

    # ================================================================
    # METADATA & CONFIDENCE
    # ================================================================

    def _detect_metadata(self, text: str, metadata: ParsedCardMetadata):
        """Detect goal, split, and other metadata from text."""
        text_lower = text.lower()

        # Detect goal
        for goal, keywords in GOAL_KEYWORDS.items():
            if any(kw in text_lower for kw in keywords):
                metadata.detected_goal = goal
                break

        # Detect split
        for split, keywords in SPLIT_KEYWORDS.items():
            if any(kw in text_lower for kw in keywords):
                metadata.detected_split = split
                break

        # Detect weeks
        weeks_match = re.search(r'(\d+)\s*(?:settiman[ae]|weeks?)', text_lower)
        if weeks_match:
            metadata.detected_weeks = int(weeks_match.group(1))

        # Also detect "4^" week pattern (Chiara's format)
        if not metadata.detected_weeks:
            weeks_match = re.search(r'(\d+)\s*[°^]\s*(?:sett)?', text_lower)
            if weeks_match:
                metadata.detected_weeks = int(weeks_match.group(1))

        # Detect sessions per week
        if not metadata.detected_sessions_per_week:
            sessions_match = re.search(
                r'(\d+)\s*(?:volte|sessioni|allenamenti|x)\s*(?:a\s*)?(?:settimana|week)',
                text_lower
            )
            if sessions_match:
                metadata.detected_sessions_per_week = int(sessions_match.group(1))

    def _calculate_confidence(self, exercises: List[ParsedExercise]) -> float:
        """Calculate overall parsing confidence."""
        if not exercises:
            return 0.0

        avg_match = sum(e.match_score for e in exercises) / len(exercises)
        has_sets = sum(1 for e in exercises if e.sets is not None) / len(exercises)
        has_reps = sum(1 for e in exercises if e.reps is not None) / len(exercises)

        # For Chiara's format, having exercises at all is a good sign
        # even without full sets/reps data
        base_confidence = min(0.3, len(exercises) * 0.03)

        confidence = base_confidence + (avg_match * 0.4) + (has_sets * 0.15) + (has_reps * 0.15)
        return round(min(confidence, 1.0), 2)
